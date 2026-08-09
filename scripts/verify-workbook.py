from pathlib import Path
import json
from openpyxl import load_workbook

root=Path(__file__).resolve().parents[1]
book=load_workbook(root/'exports/Woningencheck-Onderzoeksdatabase.xlsx',read_only=True,data_only=False)
expected=['Dashboard','Gemeenten','Regelingen','Gebiedsregels','Uitzonderingen','Aanvragen','Documenten','Bronnen','Handmatige controle','Batches','Jaarwaarden']
assert book.sheetnames==expected, f'Unexpected workbook sheets: {book.sheetnames}'
counts=json.loads((root/'exports/workbook-counts.json').read_text(encoding='utf-8'))
for name in expected[1:]:
    actual=sum(1 for row in book[name].iter_rows(values_only=True) if any(value is not None for value in row))-1
    assert actual==counts[name], f'{name}: workbook {actual}, database export {counts[name]}'
for sheet in book.worksheets:
    for row in sheet.iter_rows():
        for cell in row:
            if isinstance(cell.value,str) and cell.value.startswith(('#REF!','#DIV/0!','#VALUE!','#NAME?')):
                raise AssertionError(f'Formula error in {sheet.title}!{cell.coordinate}')
assert counts['Gemeenten']==342
print(f'PASS: canonical workbook has {len(expected)} sheets and {counts["Gemeenten"]} municipality rows')
